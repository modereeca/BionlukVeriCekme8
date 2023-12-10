import openpyxl
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup

# Excel dosyasını oluşturuyor. output_file'a her döngüde kayıt ediliyor.
output_file = "firma_bilgileri2.xlsx"
wb = Workbook()
ws = wb.active

# Url sonundaki page= kısmı önemli, aşağıda offset ona göre belirleniyor.
url = "https://www.rehberfx.com/arama?query=izmir&page="

# Ekstra veri çekmek istendiğinde, buradan başlıklar ona göre ekleniyor.
basliklar = ("FİRMA ADI", "CEP NUMARASI", "TEL", "MAİL ADRESİ", "URL")
col = 1
for baslik in basliklar:
    ws.cell(column=col, row=1).value = baslik
    col += 1

query = "#gsc.tab=0&gsc.q=izmir&gsc.page=1"
r = 2
Offset = 3959
while Offset < 6328:
    webpage_main = requests.get(url + str(Offset) + query)
    soup_main = BeautifulSoup(webpage_main.content, "html.parser")
    ilanlar = soup_main.select(".pull-left.thumbnail")
    for i in range(len(ilanlar)):
        ilan_linki = ilanlar[i].attrs.get("href")
        print(ilan_linki)

        webpage = requests.get(ilan_linki)
        soup = BeautifulSoup(webpage.content, "html.parser")

        tel = soup.select(".fa.fa-phone.fa-fw")
        mobil_tel = soup.select(".fa.fa-mobile-phone.fa-fw")
        email = soup.select(".fa.fa-envelope.fa-fw")
        firmaadi = soup.select(".media-heading.dbox-title")

        firmaadi2 = firmaadi[0].find('small').replace_with('')



        if firmaadi is not None:
            ws.cell(column=1, row=r).value = firmaadi[0].text.lstrip().rstrip()

        if len(mobil_tel) > 0:
            ws.cell(column=2, row=r).value = mobil_tel[0].parent.find_next_sibling().text

        if len(tel) > 0:
            ws.cell(column=3, row=r).value = tel[0].parent.find_next_sibling().text

        if len(email) > 0:
            ws.cell(column=4, row=r).value = email[0].parent.find_next_sibling().text

        ws.cell(column=5, row=r).value = ilan_linki
        r += 1

    Offset += 1
    print(Offset)
    wb.save(output_file)